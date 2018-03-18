# requests module is used to handle HTTP POST
import requests
# bs4 parses the HTML page
import bs4
# openpyxl handles the excel sheet formatting
import openpyxl

# to modify the URL, USN series and First subject, go to the bottom of the program
# DO NOT modify parameters unless you know what you are doing
'''
    :param sheet
        current working sheet usually a particular semester
    :param url
        url to scrub for the results
    :param
        usn is the usn to include in the HTTP post message
    :param row_no
        particular row in the excel sheet
    :param filter_sub
        subject which the crawler checks to identify valid USNs 
'''


def get_vtu_page(sheet, url, usn, row_no, filter_sub):
    form_data = {'lns': usn}
    req = requests.post(url, data=form_data)
    soup = bs4.BeautifulSoup(req.text, "html.parser")
    tags = soup.select("td")
    columns = soup.find_all("div", {"class": "divTableCell"})
    n = len(columns)
    if "1BI14CS" in usn:
        info = soup.find_all("b")
        final_marks = info[8].text[2:]
        fcd_status = info[10].text[2:]
        sheet['AA1'] = "Total Marks"
        sheet['AB1'] = "Result"
    columns.pop(0)
    columns.pop(0)
    # print(columns)
    try:
        if columns[4].getText() != filter_sub:
            return None
        sheet['C1'] = columns[4].getText()
        sheet['F1'] = columns[10].getText()
        sheet['I1'] = columns[16].getText()
        sheet['L1'] = columns[22].getText()
        sheet['O1'] = columns[28].getText()
        sheet['R1'] = columns[34].getText()
        sheet['U1'] = columns[40].getText()
        sheet['X1'] = columns[46].getText()
    except IndexError:
        return None
    cell = str(row_no)
    usn = tags[1].getText()
    name = tags[3].getText()
    sheet['A' + cell] = usn[3:]
    sheet['B' + cell] = name[1:]
    if "1BI14CS" in usn:
        sheet['AA' + cell] = final_marks
        sheet['AB' + cell] = fcd_status
    col = 'C'
    for i in range(6, n, 6):
        for j in range(3):
            sheet[col + cell] = int(columns[i + j].getText())
            col = get_next_cell(col)
        if col == "AA":
            break
            # sheet[col+cell] = columns[i+1].getText()


'''
    :param current_cell
        current cell in the excel sheet
'''


def get_next_cell(current_cell):
    # if the cell goes to Z it should shift to AA
    if current_cell == 'Z':
        return "AA"
    # python has no implicit character to ASCII value conversion so ord is required and then convert back to
    # character with chr()
    return chr(ord(current_cell) + 1)


'''
    # main function of the entire program
    :param url
        url to send the POST packet, usually the results page.php
    :param start_seq
        the initial 7 characters from which the final USN is built
    :param sheet
        the current working sheet
    :param title
        Title to be printed on the sheet
    :filter_sub
        Subject to find valid USNs
'''


def crawl(url, start_seq, sheet, title, filter_sub):
    # set the title of the current sheet
    sheet.title = title
    # format the cells  properly for each subject
    sheet.merge_cells('C1:E1')
    sheet.merge_cells('F1:H1')
    sheet.merge_cells('I1:K1')
    sheet.merge_cells('L1:N1')
    sheet.merge_cells('O1:Q1')
    sheet.merge_cells('R1:T1')
    sheet.merge_cells('U1:W1')
    sheet.merge_cells('X1:Z1')
    # print the USN and Name into the columns
    sheet['A1'] = 'USN'
    sheet['B1'] = 'Name'
    # We have to start from C but we can get_next_cell() after the for loop starts hence init_cols is B
    init_cols = 'B'
    # cell value prints from the second row
    cell_value = 2
    # loop to insert the IA,EA and Total headings in the sheet
    for i in range(0, 8):
        # move to C, D, etc
        init_cols = get_next_cell(init_cols)
        # get the cell which is a combination of the column + row
        cell = init_cols + str(cell_value)
        sheet[cell] = "IA"
        # next cell
        init_cols = get_next_cell(init_cols)
        cell = init_cols + str(cell_value)
        sheet[cell] = "EA"
        # next cell
        init_cols = get_next_cell(init_cols)
        cell = init_cols + str(cell_value)
        sheet[cell] = "Total"
    # next loop fills all the rows with extracted marks
    for i in range(1, 500):
        # print . in the output console
        print('.', end='', flush=True)
        # have to build the Roll string and pad the numbers
        # with 0
        # if integer is single digit, add 00
        # if integer is double digit, add 0
        # else do not add anything
        if len(str(i)) == 1:
            roll = "00" + str(i)
        elif len(str(i)) == 2:
            roll = "0" + str(i)
        else:
            roll = str(i)
        # build the usn
        usn = start_seq + roll
        # run the crawler
        get_vtu_page(sheet, url, usn, i + 2, filter_sub)


def calc_gpa():

    return


# open the workbook
wb = openpyxl.Workbook()
# get the first sheet, openpyxl starts sheet indexes at 0
sheet1 = wb.get_active_sheet()
# output to console
print("working", end='', flush=True)
# for the 2016 batch CS department
crawl("http://results.vtu.ac.in/vitaviresultcbcs/resultpage.php", "1BI16CS", sheet1, "3rd sem", "15MAT31")
print("\n")
# go to next sheet
sheet1 = wb.create_sheet(index=1)
# for the 2015 batch CS department
crawl("http://results.vtu.ac.in/vitaviresultcbcs/resultpage.php", "1BI15CS", sheet1, "5th sem", "15CS51")
# go to next sheet
sheet1 = wb.create_sheet(index=2)
# 2014 batch CS department
crawl("http://results.vtu.ac.in/vitaviresultnoncbcs/resultpage.php", "1BI14CS", sheet1, "7th sem", "10CS761")
print("\ndone.")
# save all modifications to the hard disk
wb.save("output.xlsx")
